import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os

# ── Configuración ──────────────────────────────────────────
API_URL = "http://localhost:3000"

BRAND_CYAN   = "FF18C5D9"
BRAND_DARK   = "FF262223"
WHITE        = "FFFFFFFF"
LIGHT_GRAY   = "FFF2F2F2"
DARK_GRAY    = "FF555555"

# ── Helpers ────────────────────────────────────────────────
def format_price(value):
    return f"$ {int(value):,}".replace(",", ".")

def payment_label(method):
    labels = {"cash": "Efectivo", "transfer": "Transferencia", "card": "Tarjeta"}
    return labels.get(method, method)

def product_label(name):
    labels = {
        "Slush": "Granizado", "DeTodito": "DeTodito",
        "Doritos": "Doritos", "Choclitos": "Choclitos",
        "Aguila Light": "Águila Light", "Pilsen": "Pilsen",
        "Water bottle": "Agua", "Syringe": "Jeringa",
        "Watermelon tape": "Cinta sandía", "Gummy": "Gomita",
        "Red Lips": "Labios rojos"
    }
    return labels.get(name, name)

def thin_border():
    side = Side(style="thin", color="FFD0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

def header_fill():
    return PatternFill("solid", fgColor=BRAND_CYAN)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Fetch data ─────────────────────────────────────────────
def fetch_report(start, end):
    url = f"{API_URL}/sales/by-date?start={start}&end={end}"
    res = requests.get(url)
    res.raise_for_status()
    return res.json()

# ── Build Excel ────────────────────────────────────────────
def build_excel(data, start, end):
    wb = openpyxl.Workbook()
    summary = data["summary"]
    sales   = data["sales"]
    slush   = data["slushBySize"]

    # ── Hoja 1: Resumen ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    # Título principal
    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = "K'llao — Reporte de Ventas"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=BRAND_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    # Período
    ws1.merge_cells("A2:D2")
    period_cell = ws1["A2"]
    period_cell.value = f"Período: {start} → {end}"
    period_cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
    period_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    period_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Espacio
    ws1.row_dimensions[3].height = 10

    # Headers de resumen
    headers = ["Métrica", "Valor"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    # Datos de resumen
    rows = [
        ("Total vendido",    format_price(summary["total"])),
        ("Número de ventas", summary["count"]),
        ("Efectivo",         format_price(summary["cash"])),
        ("Transferencia",    format_price(summary["transfer"])),
        ("Tarjeta",          format_price(summary["card"])),
    ]

    for i, (label, value) in enumerate(rows, 5):
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, val in enumerate([label, value], 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col == 1 else "right")
            cell.border = thin_border()

    # Espacio
    ws1.row_dimensions[10].height = 10

    # Granizados por tamaño
    ws1.cell(row=11, column=1, value="Granizados por tamaño").font = Font(bold=True, size=11)

    slush_headers = ["Tamaño", "Unidades vendidas"]
    for col, h in enumerate(slush_headers, 1):
        cell = ws1.cell(row=12, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    sizes_order = ["12oz", "16oz", "18oz", "22oz", "32oz"]
    for i, size in enumerate(sizes_order, 13):
        count = slush.get(size, 0)
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, val in enumerate([size, count], 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

    set_col_width(ws1, 1, 25)
    set_col_width(ws1, 2, 20)

    # ── Hoja 2: Historial ────────────────────────────────
    ws2 = wb.create_sheet("Historial de ventas")

    ws2.merge_cells("A1:F1")
    t = ws2["A1"]
    t.value = "Historial de ventas"
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=BRAND_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    sale_headers = ["Fecha", "Hora", "Productos", "Método de pago", "Total"]
    for col, h in enumerate(sale_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for i, sale in enumerate(sales, 3):
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        products_list = []
        for item in sale["items"]:
            if item["variant_id"]:
                liquor = "c/licor" if item["has_liquor"] else "s/licor"
                label = f"Granizado {item['size_name']} {liquor}"
            else:
                label = product_label(item["product_name"])
            if item["quantity"] > 1:
                label += f" ×{item['quantity']}"
            products_list.append(label)

        row_data = [
            sale["date"],
            sale["time"][:5],
            ", ".join(products_list),
            payment_label(sale["payment_method"]),
            format_price(sale["total"])
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(col == 3)
            )
            cell.border = thin_border()

    set_col_width(ws2, 1, 14)
    set_col_width(ws2, 2, 8)
    set_col_width(ws2, 3, 45)
    set_col_width(ws2, 4, 18)
    set_col_width(ws2, 5, 16)

    return wb

# ── Main ───────────────────────────────────────────────────
def main():
    if len(sys.argv) == 3:
        start = sys.argv[1]
        end   = sys.argv[2]
    else:
        today = datetime.today().strftime("%Y-%m-%d")
        start = input(f"Fecha inicio (YYYY-MM-DD) [{today}]: ").strip() or today
        end   = input(f"Fecha fin    (YYYY-MM-DD) [{today}]: ").strip() or today

    print(f"Consultando ventas del {start} al {end}...")

    data = fetch_report(start, end)

    if data["summary"]["count"] == 0:
        print("No hay ventas en ese período.")
        return

    wb = build_excel(data, start, end)

    os.makedirs("output", exist_ok=True)
    filename = f"output/reporte_{start}_{end}.xlsx"
    wb.save(filename)

    print(f"Reporte generado: {filename}")
    print(f"Ventas: {data['summary']['count']} | Total: {format_price(data['summary']['total'])}")

if __name__ == "__main__":
    main()
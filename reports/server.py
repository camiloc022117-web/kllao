from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import requests as req
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)

API_URL = "http://localhost:3000"

# ── Helpers ────────────────────────────────────────────────
def format_price(value):
    return f"$ {int(value):,}".replace(",", ".")

def payment_label(method):
    labels = {"cash": "Efectivo", "transfer": "Transferencia", "card": "Tarjeta"}
    return labels.get(method, method)

def product_label(name):
    labels = {
        "Slush": "Granizado", "DeTodito": "DeTodito",
        "Doritos": "Doritos", "Choclitos": "Choclitos",
        "Aguila Light": "Águila Light", "Pilsen": "Pilsen",
        "Water bottle": "Agua", "Syringe": "Jeringa",
        "Watermelon tape": "Cinta sandía", "Gummy": "Gomita",
        "Red Lips": "Labios rojos"
    }
    return labels.get(name, name)

def thin_border():
    side = Side(style="thin", color="FFD0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

def header_fill():
    return PatternFill("solid", fgColor="FF18C5D9")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def build_excel(data, start, end):
    BRAND_DARK = "FF262223"
    WHITE      = "FFFFFFFF"
    LIGHT_GRAY = "FFF2F2F2"
    DARK_GRAY  = "FF555555"

    wb = openpyxl.Workbook()
    summary = data["summary"]
    sales   = data["sales"]
    slush   = data["slushBySize"]

    # ── Hoja 1: Resumen ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    ws1.merge_cells("A1:D1")
    t = ws1["A1"]
    t.value = "K'llao — Reporte de Ventas"
    t.font = Font(bold=True, size=16, color=WHITE)
    t.fill = PatternFill("solid", fgColor=BRAND_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:D2")
    p = ws1["A2"]
    p.value = f"Período: {start} → {end}"
    p.font = Font(size=10, color=DARK_GRAY)
    p.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    p.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 10

    for col, h in enumerate(["Métrica", "Valor"], 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    rows = [
        ("Total vendido",    format_price(summary["total"])),
        ("Número de ventas", summary["count"]),
        ("Efectivo",         format_price(summary["cash"])),
        ("Transferencia",    format_price(summary["transfer"])),
        ("Tarjeta",          format_price(summary["card"])),
    ]

    for i, (label, value) in enumerate(rows, 5):
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, val in enumerate([label, value], 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col == 1 else "right")
            cell.border = thin_border()

    ws1.row_dimensions[10].height = 10
    ws1.cell(row=11, column=1, value="Granizados por tamaño").font = Font(bold=True, size=11)

    for col, h in enumerate(["Tamaño", "Unidades vendidas"], 1):
        cell = ws1.cell(row=12, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for i, size in enumerate(["12oz", "16oz", "18oz", "22oz", "32oz"], 13):
        count = slush.get(size, 0)
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        for col, val in enumerate([size, count], 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

    set_col_width(ws1, 1, 25)
    set_col_width(ws1, 2, 20)

    # ── Hoja 2: Historial ────────────────────────────────
    ws2 = wb.create_sheet("Historial de ventas")

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value = "Historial de ventas"
    t2.font = Font(bold=True, size=13, color=WHITE)
    t2.fill = PatternFill("solid", fgColor=BRAND_DARK)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    for col, h in enumerate(["Fecha", "Hora", "Productos", "Método de pago", "Total"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for i, sale in enumerate(sales, 3):
        fill = PatternFill("solid", fgColor=LIGHT_GRAY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        products_list = []
        for item in sale["items"]:
            if item["variant_id"]:
                liquor = "c/licor" if item["has_liquor"] else "s/licor"
                label = f"Granizado {item['size_name']} {liquor}"
            else:
                label = product_label(item["product_name"])
            if item["quantity"] > 1:
                label += f" ×{item['quantity']}"
            products_list.append(label)

        row_data = [
            sale["date"],
            sale["time"][:5],
            ", ".join(products_list),
            payment_label(sale["payment_method"]),
            format_price(sale["total"])
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(size=10)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(col == 3)
            )
            cell.border = thin_border()

    set_col_width(ws2, 1, 14)
    set_col_width(ws2, 2, 8)
    set_col_width(ws2, 3, 45)
    set_col_width(ws2, 4, 18)
    set_col_width(ws2, 5, 16)

    return wb

# ── Endpoint ───────────────────────────────────────────────
@app.route('/export')
def export():
    start = request.args.get('start')
    end   = request.args.get('end')

    if not start or not end:
        return jsonify({'error': 'start and end dates are required'}), 400

    try:
        res = req.get(f"{API_URL}/sales/by-date?start={start}&end={end}")
        res.raise_for_status()
        data = res.json()
    except Exception as e:
        return jsonify({'error': f'Error fetching data: {str(e)}'}), 500

    if data['summary']['count'] == 0:
        return jsonify({'error': 'No sales found for this period'}), 404

    wb = build_excel(data, start, end)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_{start}_{end}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(port=5000, debug=True)